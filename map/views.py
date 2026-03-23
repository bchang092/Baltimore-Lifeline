from pathlib import Path
import math

from django.conf import settings
from django.http import HttpResponse
from django.shortcuts import render


# Path to your Excel file:  BmoreLine/input_data/1109 Upload_geocoded.xlsx
XLSX_PATH = Path(settings.BASE_DIR) /"input_data" / "03232026_Upload_geocoded.xlsx"


def _to_float(x):
    """Safe float conversion; returns None if not a usable number."""
    try:
        if x is None:
            return None
        f = float(x)
        if math.isnan(f):
            return None
        return f
    except Exception:
        return None


def _load_resources_from_xlsx():
    """
    Load resources from the Excel file and return (resources_list, diagnostics_dict).

    Expected headers (case / spacing insensitive; these are what you told me earlier):
      ID
      Address
      Phone Number
      Email
      Name of Service
      Restrictions of Service
      Days of Service
      Cateogry of Help
      Description
      link to site
      Legitimate place?
      called + confirmed?
      Reliability Rate 1-10
      Call experience
      Unnamed: 18
      Latitude
      Longitude
    """

    diag = {
        "path": str(XLSX_PATH),
        "exists": XLSX_PATH.exists(),
        "sheet_title": None,
        "headers": [],
        "parsed_rows": 0,
        "skipped_no_coords": 0,
        "bad_latlng": 0,
        "errors": [],
        "sample_row": {},
        "category_header": None,
        "category_header_checked": False,
        "consolidated_categories": [],
        "tag_columns_found": [],
        "tag_counts": {},
        "taglist_header": None,
    }

    resources = []

    if not XLSX_PATH.exists():
        diag["errors"].append("File not found")
        return resources, diag

    try:
        from openpyxl import load_workbook
    except ImportError:
        diag["errors"].append("openpyxl not installed (pip install openpyxl)")
        return resources, diag

    try:
        wb = load_workbook(filename=str(XLSX_PATH), data_only=True)
        ws = wb.active
        diag["sheet_title"] = ws.title

        # Header row
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [h or "" for h in header_row]
        diag["headers"] = headers
        print("Detected headers:", headers)

        def norm(s):
            return " ".join(str(s).strip().lower().split()) if s is not None else ""

        header_map = {norm(h): idx for idx, h in enumerate(headers)}

        def grab(row, *names, default=""):
            """Pull a cell by any of the given header names."""
            for n in names:
                idx = header_map.get(norm(n))
                if idx is not None and idx < len(row):
                    val = row[idx]
                    if val is not None:
                        return val
            return default

        def first_matching_header(*names):
            for n in names:
                if norm(n) in header_map:
                    return n
            return None

        def to_bool_flag(v):
            if v is None:
                return None
            s = str(v).strip().lower()
            if s in {"yes", "y", "true", "1"}:
                return True
            if s in {"no", "n", "false", "0"}:
                return False
            return None  # unknown / not filled

        def is_truthy_tag(v):
            if v is None:
                return False
            if isinstance(v, bool):
                return v
            if isinstance(v, (int, float)):
                return v != 0 and not math.isnan(v)
            s = str(v).strip().lower()
            if s in {"yes", "y", "true", "1", "x"}:
                return True
            if s in {"no", "n", "false", "0", ""}:
                return False
            return False

        def parse_tag_list(val):
            if val is None:
                return []
            s = str(val).strip()
            if not s:
                return []
            parts = []
            for chunk in s.replace("\n", ",").replace(";", ",").replace("|", ",").split(","):
                tag = chunk.strip()
                if tag:
                    parts.append(tag)
            return parts

        category_header = first_matching_header(
            "Category of Help",
            "Cateogry of Help",
            "Cateogry of Help (Original)",
            "Category",
        )
        if category_header:
            diag["category_header"] = category_header
            print("Category column selected:", category_header)
        else:
            print("Category column selected:", False)

        consolidated_categories = set()
        tag_columns = [f"Tag_{i:02d}" for i in range(1, 26)]
        tag_columns_found = [c for c in tag_columns if norm(c) in header_map]
        diag["tag_columns_found"] = tag_columns_found
        if tag_columns_found:
            print("Tag columns found:", tag_columns_found)
        else:
            print("Tag columns found:", False)
        tag_counts = {c: 0 for c in tag_columns_found}

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(grab(row, "Name of Service", "Name", default="")).strip()
            address = str(grab(row, "Address", default="")).strip()
            phone = str(grab(row, "Phone Number", "Phone", default="")).strip()
            email = str(grab(row, "Email", default="")).strip()
            category = str(
                grab(
                    row,
                    "Category of Help",
                    "Cateogry of Help",
                    "Cateogry of Help (Original)",
                    "Category",
                    default="",
                )
            ).strip() or "Uncategorized"
            consolidated_value = str(
                grab(
                    row,
                    "Consolidated Category",
                    "Consolidated Tag Category",
                    default="",
                )
            ).strip()
            if consolidated_value:
                consolidated_categories.add(consolidated_value)
            taglist_header = first_matching_header("Tag_List", "Taglist", "Tag List")
            if taglist_header and diag.get("taglist_header") is None:
                diag["taglist_header"] = taglist_header
                print("Tag list column selected:", taglist_header)

            taglist_value = grab(row, "Tag_List", "Taglist", "Tag List", default="")
            tags = parse_tag_list(taglist_value)
            if not tags:
                for col in tag_columns_found:
                    val = grab(row, col, default=None)
                    if is_truthy_tag(val):
                        tags.append(col)
                        tag_counts[col] += 1
            desc = str(grab(row, "Description", default="")).strip()
            restrictions = str(grab(row, "Restrictions of Service", default="")).strip()
            days = str(grab(row, "Days of Service", default="")).strip()
            link = str(
                grab(row, "link to site", "Website", "Link", default="")
            ).strip()

            legit_raw = grab(row, "Legitimate place?", "Legitimate place ?", default="")
            confirmed_raw = grab(
                row,
                "confirmed",
                "called + confirmed?",
                "called + confirmed ?",
                default="",
            )

            reliability_raw = str(
                grab(row, "Reliability Rate 1-10", "Reliability Rate 1–10", "Reliability", default="")
            ).strip()
            avg_reliability_ratings_raw = str(
                grab(
                    row,
                    "avg_reliability_ratings",
                    "Average Reliability Ratings",
                    default="",
                )
            ).strip()
            avg_reliability_ratings = (
                avg_reliability_ratings_raw
                if avg_reliability_ratings_raw.lower() not in {"", "nan", "none"}
                else "na"
            )
            condensed_reliability_description = str(
                grab(row, "Condensed Reliability Description", default="")
            ).strip()
            reliability = (
                avg_reliability_ratings
                if avg_reliability_ratings != "na"
                else reliability_raw if reliability_raw not in {"", "nan", "none"} else "na"
            )

            call_exp = str(grab(row, "Call experience", default="")).strip()
            extra = str(grab(row, "Unnamed: 18", default="")).strip()
            call_notes = " | ".join([x for x in [call_exp, extra] if x])

            lat_raw = grab(row, "Latitude", "Lat", default=None)
            lng_raw = grab(row, "Longitude", "Lng", "Long", default=None)

            lat = _to_float(lat_raw)
            lng = _to_float(lng_raw)

            # Skip rows without usable coordinates
            if lat is None or lng is None:
                diag["skipped_no_coords"] += 1
                continue

            # Filter obviously invalid coordinates
            if not (-90 <= lat <= 90 and -180 <= lng <= 180):
                diag["bad_latlng"] += 1
                continue

            rid = grab(row, "ID", "id", default=None)
            if rid is None or str(rid).strip() == "":
                rid = len(resources) + 1

            res = {
                "id": rid,
                "name": name or "Unnamed resource",
                "lat": lat,
                "lng": lng,
                "category": category,
                "phone_number": phone,
                "address": address,
                "email": email,
                "description": desc,
                "restrictions": restrictions,
                "days": days,
                "link": link,
                "legit": to_bool_flag(legit_raw),
                "confirmed": to_bool_flag(confirmed_raw),
                "reliability": reliability,
                "avg_reliability_ratings": avg_reliability_ratings,
                "condensed_reliability_description": condensed_reliability_description,
                "call_notes": call_notes,
                "tags": tags,
            }
            resources.append(res)

        diag["parsed_rows"] = len(resources)
        diag["consolidated_categories"] = sorted(consolidated_categories)
        diag["tag_counts"] = tag_counts
        if resources:
            diag["sample_row"] = resources[0]

    except Exception as e:
        diag["errors"].append(f"{type(e).__name__}: {e}")

    return resources, diag


def resources_map(request):
    """Main map view – loads data from Excel and passes it into the template."""
    resources, diag = _load_resources_from_xlsx()
    categories = sorted({r.get("category") for r in resources if r.get("category")})
    print("Categories passed to template:", categories)
    consolidated_categories = diag.get("consolidated_categories", [])

    # Quick debug view: /resources_map/?debug=1
    if request.GET.get("debug") == "1":
        return HttpResponse(
            f"DEBUG – Excel path: {diag['path']}\n"
            f"Exists: {diag['exists']}\n"
            f"Sheet: {diag['sheet_title']}\n"
            f"Headers: {diag['headers']}\n"
            f"Parsed rows (with coords): {diag['parsed_rows']}\n"
            f"Skipped (no coords): {diag['skipped_no_coords']}\n"
            f"Bad lat/lng: {diag['bad_latlng']}\n"
            f"Tag columns found: {diag.get('tag_columns_found')}\n"
            f"Tag counts: {diag.get('tag_counts')}\n"
            f"Errors: {diag['errors']}\n"
            f"Sample row: {diag['sample_row']}"
            .replace("\n", "<br>")
        )

    # Normal map render
    return render(
        request,
        "map_home.html",
        {"resources": resources, "consolidated_categories": consolidated_categories},
    )


def _derive_tags_from_answers(answers):
    need = set()
    scenario = set()
    barrier = set()
    benefit = set()
    insurance = set()
    housing_status = set()
    risk = set()
    demo = set()
    pathway = set()
    access = set()
    doc = set()
    mobility = set()

    def add(target, *tags):
        for tag in tags:
            if tag:
                target.add(tag)

    a1 = answers.get("A1_safe_tonight")
    a2 = answers.get("A2_sleep_tonight")
    a3 = answers.get("A3_threats_abuse")
    a4 = set(answers.get("A4_needs_today", []))
    b1 = answers.get("B1_where_staying")
    b2 = answers.get("B2_losing_housing")
    b3 = set(answers.get("B3_utilities", []))
    c1 = answers.get("C1_food_2_3_days")
    c2 = set(answers.get("C2_needs", []))
    d1 = answers.get("D1_income")
    d2 = set(answers.get("D2_trouble_paying", []))
    d3 = set(answers.get("D3_benefits", []))
    d4 = answers.get("D4_lost_medicaid")
    d5 = set(answers.get("D5_help_applying", []))
    e1 = set(answers.get("E1_health_needs", []))
    e2 = answers.get("E2_have_doctor")
    f1 = answers.get("F1_transport")
    f2 = answers.get("F2_phone")
    g1 = set(answers.get("G1_documents", []))
    g2 = set(answers.get("G2_barriers", []))
    h1 = set(answers.get("H1_household", []))
    i1 = answers.get("I1_output_preference")

    # Crisis / tonight rules
    if a1 == "no" or a3 == "yes":
        add(risk, "R003_DOMESTIC_VIOLENCE_IMMEDIATE")
        add(scenario, "S005_FLEEING_VIOLENCE")
        add(barrier, "B019_DOMESTIC_VIOLENCE")
        add(housing_status, "H015_FLEEING_DV")
        add(pathway, "P001_CRISIS_TONIGHT")

    if a2 == "no" or b1 in {"outside", "car", "abandoned"}:
        add(risk, "R002_NO_SHELTER_TONIGHT")
        add(scenario, "S004_HOMELESS_TONIGHT")
        add(need, "N010_SHELTER_TONIGHT")
        add(pathway, "P002_SHELTER_TONIGHT")
        if b1 == "outside":
            add(housing_status, "H005_STREET_HOMELESS")
        elif b1 == "car":
            add(housing_status, "H006_CAR_HOMELESS")
        elif b1 == "abandoned":
            add(housing_status, "H007_ABANDONED_BUILDING")

    if c1 == "no":
        add(risk, "R001_NO_FOOD_NEXT_24H")
        add(scenario, "S001_NO_FOOD")
        add(need, "N002_GROCERIES")
        add(pathway, "P003_FOOD_TODAY")
    elif c1 == "sometimes":
        add(need, "N002_GROCERIES")

    # Housing status from B1
    if b1 == "own_place":
        add(housing_status, "H001_HOUSED_STABLE")
    elif b1 == "friends_family":
        add(housing_status, "H003_TEMPORARY_COUCH", "H009_DOUBLING_UP")
        add(scenario, "S006_COUCH_SURFING")
    elif b1 == "shelter":
        add(housing_status, "H004_SHELTER")
    elif b1 == "hotel_motel":
        add(housing_status, "H008_HOTEL_MOTEL")

    # Housing stability
    if b2 == "yes" and b1 in {"own_place", "friends_family"}:
        add(housing_status, "H002_HOUSED_AT_RISK", "H010_AT_RISK_30_DAYS")
        add(scenario, "S007_EVICTION_NOTICE")
        add(pathway, "P016_HOUSING_STABILITY_PATHWAY")

    if "rent_utilities" in a4 or "rent_mortgage" in d2:
        add(need, "N015_RENT_ASSISTANCE")
        add(scenario, "S008_RENT_BEHIND")
        add(pathway, "P007_RENT_RELIEF_THIS_WEEK")

    # A4 needs today mapping
    if "food" in a4:
        add(need, "N002_GROCERIES", "N003_HOT_MEAL")
        add(scenario, "S001_NO_FOOD")
        add(pathway, "P003_FOOD_TODAY")
    if "sleep" in a4:
        add(need, "N010_SHELTER_TONIGHT")
        add(scenario, "S004_HOMELESS_TONIGHT")
        add(pathway, "P002_SHELTER_TONIGHT")
    if "clothes_hygiene" in a4:
        add(need, "N020_CLOTHES", "N023_HYGIENE_KITS")
    if "phone_internet" in a4:
        add(need, "N040_PHONE_SERVICE", "N042_INTERNET_ACCESS")
        add(barrier, "B005_NO_PHONE", "B007_NO_INTERNET")
        add(pathway, "P018_PHONE_ACCESS_PATHWAY")
    if "transportation" in a4:
        add(need, "N030_TRANSIT_PASS", "N031_RIDES_TO_APPTS")
        add(barrier, "B017_NO_TRANSPORT")
    if "medical_mental" in a4:
        add(need, "N050_PRIMARY_CARE", "N051_MENTAL_HEALTH")
        add(pathway, "P009_HEALTHCARE_ESTABLISH_PCP", "P010_MENTAL_HEALTH_PATHWAY")
    if "medicaid_insurance" in a4:
        add(insurance, "I006_UNSURE_INSURANCE")
        add(pathway, "P005_MEDICAID_HELP_THIS_WEEK")
    if "benefits_money" in a4:
        add(benefit, "F008_NO_BENEFITS", "F013_CONFUSED_BENEFITS")
        add(pathway, "P004_BENEFITS_THIS_WEEK")
    if "childcare_family" in a4:
        add(demo, "D007_PARENT", "D008_CAREGIVER")
        add(pathway, "P012_CHILDCARE_PATHWAY")
    if "disability_long_term" in a4:
        add(demo, "D013_PERSON_WITH_DISABILITY")
        add(barrier, "B008_DISABILITY_MOBILITY", "B009_DISABILITY_COGNITIVE")
        add(pathway, "P020_LONG_TERM_SUPPORTS")
    if "education_job" in a4:
        add(pathway, "P013_JOB_PATHWAY", "P014_EDUCATION_PATHWAY")
    if "legal" in a4:
        add(pathway, "P015_LEGAL_ASSISTANCE_PATHWAY")

    if d1 in {"no", "sometimes"}:
        add(benefit, "F013_CONFUSED_BENEFITS")
        add(pathway, "P004_BENEFITS_THIS_WEEK")

    # Utilities
    if b3.intersection({"electric", "gas", "water"}):
        add(need, "N016_UTILITY_ASSISTANCE")
        add(pathway, "P008_UTILITY_RELIEF_THIS_WEEK")
        if "electric" in b3:
            add(scenario, "S012_NO_POWER")
        if "gas" in b3:
            add(scenario, "S011_NO_HEAT")
        if "water" in b3:
            add(need, "N017_WATER_ASSISTANCE")
            add(scenario, "S010_NO_WATER")

    if "internet_phone" in b3 or f2 == "no" or "phone_internet" in a4:
        add(need, "N040_PHONE_SERVICE", "N042_INTERNET_ACCESS")
        if f2 == "no":
            add(scenario, "S018_NO_PHONE")
        elif f2 == "sometimes":
            add(scenario, "S019_NO_DATA_LIMITED")
        add(barrier, "B005_NO_PHONE", "B007_NO_INTERNET")
        add(pathway, "P018_PHONE_ACCESS_PATHWAY")

    # Food
    if "hot_meals" in c2:
        add(scenario, "S003_HOT_MEAL_NEEDED")
        add(need, "N003_HOT_MEAL")
        add(pathway, "P003_FOOD_TODAY")

    if "baby_food" in c2:
        add(need, "N005_BABY_FORMULA")
        add(scenario, "S015_CHILDREN_NO_FOOD")
        add(demo, "D007_PARENT")

    # Essentials
    if "hygiene" in c2 or "clothes_hygiene" in a4:
        add(need, "N023_HYGIENE_KITS")

    if "clothing_work" in c2:
        add(need, "N022_WORK_CLOTHES")
        add(scenario, "S017_WORK_CLOTHES_NEEDED")

    if "clothing_everyday" in c2:
        add(need, "N020_CLOTHES")

    if "winter_clothing" in c2:
        add(need, "N021_WARM_CLOTHES", "N024_BEDDING")
        if b1 in {"outside", "car"}:
            add(risk, "R007_OUTSIDE_IN_COLD")

    # Benefits and insurance
    if "none" in d3 and (d5 or "not_sure" in d5):
        add(benefit, "F008_NO_BENEFITS", "F013_CONFUSED_BENEFITS")
        add(pathway, "P004_BENEFITS_THIS_WEEK")

    if "snap" in d3:
        add(benefit, "F001_HAS_SNAP")
    if "medicaid" in d3:
        add(benefit, "F003_HAS_MEDICAID")
    if "medicare" in d3:
        add(benefit, "F004_HAS_MEDICARE")
    elif "snap" in d5 or ("food" in d2 and "snap" not in d3):
        add(benefit, "F016_SNAP_PENDING", "F013_CONFUSED_BENEFITS")
        add(pathway, "P006_SNAP_THIS_WEEK")

    if d4 == "yes":
        add(scenario, "S029_LOST_MEDICAID")
        add(benefit, "F011_LOST_MEDICAID")
        add(insurance, "I007_LOST_MEDICAID_RED")
        add(pathway, "P005_MEDICAID_HELP_THIS_WEEK")

    if "medicaid" in d5:
        add(insurance, "I008_MEDICAID_NEEDS_RENEWAL", "I006_UNSURE_INSURANCE")
        add(pathway, "P005_MEDICAID_HELP_THIS_WEEK")
    if "medicare" in d5:
        add(insurance, "I006_UNSURE_INSURANCE")
        add(pathway, "P005_MEDICAID_HELP_THIS_WEEK")
    if "ssi_ssdi" in d5 or "unemployment" in d5 or "childcare" in d5:
        add(benefit, "F013_CONFUSED_BENEFITS")
        add(pathway, "P004_BENEFITS_THIS_WEEK")

    if e2 == "no" and (("medical_mental" in a4) or e1):
        add(insurance, "I001_NO_INSURANCE")
        add(insurance, "I014_NO_PCP")
        add(scenario, "S025_NO_PRIMARY_CARE")
        add(pathway, "P009_HEALTHCARE_ESTABLISH_PCP")

    # Healthcare
    if "primary_care" in e1 or e2 == "no":
        add(need, "N050_PRIMARY_CARE")
        add(scenario, "S025_NO_PRIMARY_CARE")
        add(insurance, "I014_NO_PCP")
        add(pathway, "P009_HEALTHCARE_ESTABLISH_PCP")

    if "mental_health" in e1:
        add(need, "N051_MENTAL_HEALTH")
        add(scenario, "S026_NO_MENTAL_HEALTH")
        add(barrier, "B014_MENTAL_HEALTH_BARRIER")
        add(pathway, "P010_MENTAL_HEALTH_PATHWAY")

    if "substance_use" in e1:
        add(need, "N052_SUBSTANCE_USE")
        add(scenario, "S027_RECOVERY_SUPPORT")
        add(barrier, "B015_ADDICTION_BARRIER")
        add(pathway, "P011_SUBSTANCE_PATHWAY")

    if "prescriptions" in e1 or "medical" in d2:
        add(need, "N057_PHARMACY_ASSIST")
        add(insurance, "I013_MEDICATION_COST_ISSUE")

    if "dental" in e1:
        add(need, "N053_DENTAL")
    if "vision" in e1:
        add(need, "N054_VISION")
    if "equipment" in e1:
        add(need, "N056_MED_EQUIPMENT")
    if "prenatal" in e1:
        add(need, "N055_PRENATAL")
        add(insurance, "I018_PRENATAL_CARE_NEEDED")
        add(demo, "D006_PREGNANT")

    # Transportation
    if f1 == "no_reliable" and e1:
        add(barrier, "B017_NO_TRANSPORT")
        add(need, "N031_RIDES_TO_APPTS")
        add(scenario, "S022_NEED_TRANSIT_TO_MEDICAL")

    food_insecure = c1 in {"no", "sometimes"} or "food" in a4
    if (f1 == "no_reliable" or "transportation" in a4) and food_insecure:
        add(need, "N034_RIDES_TO_PANTRY")
        add(scenario, "S023_NEED_TRANSIT_TO_PANTRY")

    # Documents and barriers
    if "none" in g1:
        add(barrier, "B001_NO_ID", "B002_NO_SSN", "B003_NO_BIRTH_CERT", "B004_NO_DOCUMENTS")
        add(doc, "DOC002_NO_ID", "DOC004_NO_SSN", "DOC006_NO_BC")
        add(pathway, "P017_ID_RECOVERY_PATHWAY")

    if "immigration" in g2:
        add(barrier, "B013_IMMIGRATION_FEAR", "B038_UNDOCUMENTED")
        add(demo, "D010_UNDOCUMENTED")

    if "disability" in g2:
        add(barrier, "B008_DISABILITY_MOBILITY", "B009_DISABILITY_COGNITIVE")
        add(demo, "D013_PERSON_WITH_DISABILITY")

    if "criminal_record" in g2:
        add(barrier, "B012_CRIMINAL_RECORD")
        add(demo, "D012_REENTRY")

    if "limited_english" in g2:
        add(barrier, "B010_LANGUAGE_BARRIER", "B011_LOW_LITERACY")
        add(access, "A008_LOW_LITERACY_MODE")

    if "no_phone_internet" in g2:
        add(barrier, "B005_NO_PHONE", "B007_NO_INTERNET", "B024_TECH_BARRIER")
        add(need, "N040_PHONE_SERVICE", "N042_INTERNET_ACCESS")

    # Household
    if "children" in h1:
        add(demo, "D007_PARENT")
    if "older_adult" in h1:
        add(demo, "D005_OLDER_ADULT", "D008_CAREGIVER")
    if "disability" in h1:
        add(demo, "D013_PERSON_WITH_DISABILITY")

    # Output preferences
    if i1 == "text":
        add(mobility, "M007_PREFER_TEXT")
    if i1 == "email":
        add(mobility, "M008_PREFER_EMAIL")
    if i1 == "printable":
        add(mobility, "M009_PREFER_PRINT")
    if i1 == "helper":
        add(access, "A010_SIMPLE_MODE")

    # Cleanup rules
    if "F001_HAS_SNAP" in benefit and "F008_NO_BENEFITS" in benefit:
        benefit.discard("F008_NO_BENEFITS")
    if ("F003_HAS_MEDICAID" in benefit or "F004_HAS_MEDICARE" in benefit) and "I001_NO_INSURANCE" in insurance:
        insurance.discard("I001_NO_INSURANCE")

    all_tags = set().union(
        need,
        scenario,
        barrier,
        benefit,
        insurance,
        housing_status,
        risk,
        demo,
        pathway,
        access,
        doc,
        mobility,
    )

    return {
        "crisis_tags": sorted(risk),
        "need_tags": sorted(need),
        "scenario_tags": sorted(scenario),
        "barrier_tags": sorted(barrier),
        "benefit_tags": sorted(benefit),
        "insurance_tags": sorted(insurance),
        "housing_status_tags": sorted(housing_status),
        "demographic_tags": sorted(demo),
        "pathway_tags": sorted(pathway),
        "all_tags_deduped": sorted(all_tags),
    }


def home_page(request):
    return render(request, "home.html")


def questionnaire_page(request):
    if request.method == "POST":
        def get(name):
            return (request.POST.get(name) or "").strip()

        def get_list(name):
            return [v for v in request.POST.getlist(name) if v]

        answers = {
            "A1_safe_tonight": get("A1_safe_tonight"),
            "A2_sleep_tonight": get("A2_sleep_tonight"),
            "A3_threats_abuse": get("A3_threats_abuse"),
            "A4_needs_today": get_list("A4_needs_today"),
            "B1_where_staying": get("B1_where_staying"),
            "B2_losing_housing": get("B2_losing_housing"),
            "B3_utilities": get_list("B3_utilities"),
            "C1_food_2_3_days": get("C1_food_2_3_days"),
            "C2_needs": get_list("C2_needs"),
            "D1_income": get("D1_income"),
            "D2_trouble_paying": get_list("D2_trouble_paying"),
            "D3_benefits": get_list("D3_benefits"),
            "D4_lost_medicaid": get("D4_lost_medicaid"),
            "D5_help_applying": get_list("D5_help_applying"),
            "E1_health_needs": get_list("E1_health_needs"),
            "E2_have_doctor": get("E2_have_doctor"),
            "F1_transport": get("F1_transport"),
            "F2_phone": get("F2_phone"),
            "G1_documents": get_list("G1_documents"),
            "G2_barriers": get_list("G2_barriers"),
            "H1_household": get_list("H1_household"),
            "I1_output_preference": get("I1_output_preference"),
        }

        tags = _derive_tags_from_answers(answers)
        tagset = set(tags["all_tags_deduped"])

        resources, _diag = _load_resources_from_xlsx()

        if tagset:
            filtered = []
            for r in resources:
                r_tags = set(r.get("tags") or [])
                if r_tags & tagset:
                    filtered.append(r)
        else:
            filtered = resources

        filtered.sort(key=lambda r: len(set(r.get("tags") or []) & tagset), reverse=True)

        return render(
            request,
            "map_recommended.html",
            {
                "resources": filtered,
                "tags": tags,
                "tag_count": len(tagset),
            },
        )

    return render(request, "questionnaire.html")


def actions_page(request):
    return render(request, "actions.html")


def about_page(request):
    return render(request, "about.html")


# debug endpoint still available
def ping(request):
    return HttpResponse("pong")
